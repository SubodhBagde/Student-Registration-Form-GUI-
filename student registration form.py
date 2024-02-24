from tkinter import *
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image, ImageTk
import os
from tkinter.ttk import Combobox
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib

background = "white"
framebg = "white"
framefg = "white"

root = Tk()
root.title("Student Registration Form")
root.geometry("1250x700+210+100")
root.config(bg=background)

file = pathlib.Path('Student_data.xlsx')
if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    sheet['A1'] = "Registration Number"
    sheet['B1'] = "Date"
    sheet['C1'] = "Full Name (First)"
    sheet['D1'] = "Full Name (Middle)"
    sheet['E1'] = "Full Name (Last)"
    sheet['F1'] = "Student Number"
    sheet['G1'] = "Year Level"
    sheet['H1'] = "Degree Program"
    sheet['I1'] = "Email"
    sheet['J1'] = "Birth Date (Month)"
    sheet['K1'] = "Birth Date (Day)"
    sheet['L1'] = "Birth Date (Year)"
    sheet['M1'] = "Present Address"
    sheet['N1'] = "Gender"

    file.save('Student_data.xlsx')
    
# Exit
def Exit():
    root.destroy()

# Show image
def Showimage():
    global filename
    global img
    filename = filedialog.askopenfilename(initialdir=os.getcwd(), title="Select image file", filetypes=(("JPG files","*.jpg"),("PNG file", "*.png"),("All files", "*.txt")))

    img = (Image.open(filename))
    resized_image = img.resize((190,190))
    photo = ImageTk.PhotoImage(resized_image)
    p.config(image=photo)
    p.image=photo

# Clear
def Clear():
    global img
    Registration.set('')
    Name1.set('')
    Name2.set('')
    Name3.set('')
    Phone.set('')
    Year.set('Please Select')
    Degree.set('Please Select')
    Email.set('')
    DOB1.set('Select a Month')
    DOB2.set('Select a Day')
    DOB3.set('Select a Year')
    Address.set('')
    SaveButton.config(state='normal')
    img1 = PhotoImage(file='Images/user1.png')
    p.config(image=img1)
    p.image=img1
    img=""

# Save
def Save():
    R1 = Registration.get()
    D = Date.get()
    N1 = Name1.get()
    N2 = Name2.get()
    N3 = Name3.get()
    Y1 = Year.get()
    try:
        g1 = gender
    except:
        messagebox.showerror("error", "Please select Gender!")

    P1 = Phone.get()
    D1 = DOB1.get()
    D2 = DOB2.get()
    D3 = DOB3.get()
    Deg = Degree.get()
    E1 = Email.get()
    A1 = Address.get()

    if (R1=="" or N1=="" or N2=="" or N3=="" or Y1=="Please Select" or P1=="" or D1=="Select a Month" or D2=="Select a Day" or D3=="Select a Year" or Deg=="Please Select" or E1=="" or A1==""):
        messagebox.showerror("error", "Few details are missing!")
    else:
        file = openpyxl.load_workbook('Student_data.xlsx')
        sheet = file.active
        sheet.cell(column=1, row=sheet.max_row+1, value=R1)
        sheet.cell(column=2, row=sheet.max_row, value=D)
        sheet.cell(column=3, row=sheet.max_row, value=N1)
        sheet.cell(column=4, row=sheet.max_row, value=N2)
        sheet.cell(column=5, row=sheet.max_row, value=N3)
        sheet.cell(column=6, row=sheet.max_row, value=P1)
        sheet.cell(column=7, row=sheet.max_row, value=Y1)
        sheet.cell(column=8, row=sheet.max_row, value=Deg)
        sheet.cell(column=9, row=sheet.max_row, value=E1)
        sheet.cell(column=10, row=sheet.max_row, value=D1)
        sheet.cell(column=11, row=sheet.max_row, value=D2)
        sheet.cell(column=12, row=sheet.max_row, value=D3)
        sheet.cell(column=13, row=sheet.max_row, value=A1)
        sheet.cell(column=14, row=sheet.max_row, value=g1)
        file.save(r'Student_data.xlsx')

        try:
            img.save("Student_images/"+str(R1)+".jpg")
        except:
            messagebox.showinfo("info", "Profile picture is not available!")
        messagebox.showinfo("info", "Successfully data entered!")
        Clear()

# Search
# def search():
#     text = Search.get()
#     Clear()
#     SaveButton.config(state='disabled')
#     file = openpyxl.load_workbook("Student_data.xlsx")
#     sheet = file.active
#     for row in sheet.rows:
#         if row[0].value == int(text):
#             name = row[0]
#             print(str(name))
            # reg_no_position = str(name)[14:-1]
            # reg_number = str(name)[15:-1]
            # print(reg_no_position)
            # print(reg_number)
    
    # try:
    #     print(str(name))
    # except:
    #     messagebox.showerror("Invalid", "Invalid registration number!")
    
    # x1 = sheet.cell(row=int(reg_number), column=1).value
    # x2 = sheet.cell(row=int(reg_number), column=2).value
    # x3 = sheet.cell(row=int(reg_number), column=3).value
    # x4 = sheet.cell(row=int(reg_number), column=4).value
    # x5 = sheet.cell(row=int(reg_number), column=5).value
    # x6 = sheet.cell(row=int(reg_number), column=6).value
    # x7 = sheet.cell(row=int(reg_number), column=7).value
    # x8 = sheet.cell(row=int(reg_number), column=8).value
    # x9 = sheet.cell(row=int(reg_number), column=9).value
    # x10 = sheet.cell(row=int(reg_number), column=10).value
    # x11 = sheet.cell(row=int(reg_number), column=11).value
    # x12 = sheet.cell(row=int(reg_number), column=12).value
    # x13 = sheet.cell(row=int(reg_number), column=13).value
    # x14 = sheet.cell(row=int(reg_number), column=14).value

    # print(x1)
    # print(x2)
    # print(x3)
    # print(x4)
    # print(x5)
    # print(x6)
    # print(x7)
    # print(x8)
    # print(x9)
    # print(x10)
    # print(x11)
    # print(x12)
    # print(x13)
    # print(x14)

    # Registration.set(x1)
    # Date.set(x2)
    # Name1.set(x3)
    # Name2.set(x4)
    # Name3.set(x5)
    # Phone.set(x6)
    # Year.set(x7)
    # Degree.set(x8)
    # Email.set(x9)
    # DOB1.set(x10)
    # DOB2.set(x11)
    # DOB3.set(x12)
    # Address.set(x13)
    # if x14 == "Male":
    #     G1.select()
    # else:
    #     G2.select()

# Gender
def selection():
    global gender
    value = Gender.get()
    if value==1:
        gender="Male"
    else:
        gender="Female"

# Top Frames
Label(root,text="G H Raisoni College of Engineering Nagpur", width=2, font="Arial 16 bold", fg="#C0392B", height=1, bg="white").pack(side=TOP, fill=X)
Label(root,text="Accredited “A++” by NAAC in Third Cycle", width=2, font="Montserrat 14 bold", fg="black", height=1, bg="white").pack(side=TOP, fill=X)
Label(root,text="Empowered Autonomous Institution Affiliated to Rashtrasant Tukadoji Maharaj Nagpur University, Nagpur", width=2, font="Montserrat 14 bold", fg="black", height=1, bg="white").pack(side=TOP, fill=X)
Label(root,text="DTE Code : 4116", width=2, font="Arial 16 bold", fg="black", height=1, bg="white").pack(side=TOP, fill=X)

Label(root,text="STUDENT REGISTRATION FORM", width=2, font="Montserrat 20 bold", fg="white", height=2, bg="#154360").pack(side=TOP, fill=X)

# Search box to update
Search = StringVar()
search_box = Entry(root, textvariable=Search, width=13, bd=2, font="montserrat 16")
search_box.place(x=1200,y=135)
search_icon = PhotoImage(file="Images/search.png")
search_button = Button(root,compound=LEFT, height=25, image=search_icon, width=30, bg="white")
search_button.place(x=1365, y=135)
update_icon = PhotoImage(file="Images/update.png")
update_button = Button(root, image=update_icon)
update_button.place(x=300, y=135)
Label(root, text="UPDATE", font="Montserrat 7 bold", fg="white", bg="#154360" ).place(x=292, y=165)

# Canvas
c = Canvas(root, bg="lightgrey", height=700, width=1550)
c.pack()

# Date & Registration Number
Label(root, text="Registration Number ", font="Montserrat 14 bold", fg="#212F3C", bg="lightgrey").place(x=50, y=200)
Registration = StringVar()
Registration_box = Entry(root, textvariable=Registration, width=17, bd=2, relief="sunken", font="Montserrat 12")
Registration_box.place(x=254, y=202)
Label(root, text="Date ", font="Montserrat 14 bold", fg="#212F3C", bg="lightgrey").place(x=500, y=200)
Date = StringVar()
today = date.today()
d1 = today.strftime("%d/%m/%y")
Date_box = Entry(root, textvariable=Date, state="disabled", width=13, bd=2, relief="sunken", font="Montserrat 12")
Date_box.place(x=557, y=202)
Date.set(d1)

# Sudent Details
obj = LabelFrame(root, text="Student Details", bg="white", font="comicsansms 18 bold", bd=2, width=1150, height=500, relief="groove", fg="#212F3C")
obj.place(x=40, y=250)

Label(root, text="Full Name", font="Montserrat 16 bold", fg="#212F3C", bg="white").place(x=95, y=300)
Label(root, text="First Name", font="Montserrat 10 ", fg="#515A5A", bg="white").place(x=95, y=372)
Label(root, text="Middle Name", font="Montserrat 10 ", fg="#515A5A", bg="white").place(x=265, y=375)
Label(root, text="Last Name", font="Montserrat 10 ", fg="#515A5A", bg="white").place(x=435, y=374)
Label(root, text="Student Phone Number", font="Montserrat 16 bold", fg="#212F3C", bg="white").place(x=695, y=300)
Label(root, text="eg: 999998888", font="Montserrat 10 ", fg="#515A5A", bg="white").place(x=695, y=370)
Label(root, text="Year Level", font="Montserrat 16 bold", fg="#212F3C", bg="white").place(x=695, y=425)
Label(root, text="Degree Program", font="Montserrat 16 bold", fg="#212F3C", bg="white").place(x=95, y=425)
Label(root, text="Email", font="Montserrat 16 bold", fg="#212F3C", bg="white").place(x=95, y=535)
Label(root, text="eg: name@example.com", font="Montserrat 10 ", fg="#515A5A", bg="white").place(x=95, y=610)
Label(root, text="Date of Birth", font="Montserrat 16 bold", fg="#212F3C", bg="white").place(x=695, y=535)
Label(root, text="Present Address", font="Montserrat 16 bold", fg="#212F3C", bg="white").place(x=95, y=655)
Label(root, text="Gender", font="Montserrat 16 bold", fg="#212F3C", bg="white").place(x=695, y=655)

Name1 = StringVar()
Name2 = StringVar()
Name3 = StringVar()
Name_box1 = Entry(root, textvariable=Name1, width=14, bd=2, relief="sunken", font="Montserrat 14", fg="#424949")
Name_box1.place(x=95, y=344)
Name_box2 = Entry(root, textvariable=Name2, width=14, bd=2, relief="sunken", font="Montserrat 14", fg="#424949")
Name_box2.place(x=265, y=344)
Name_box3 = Entry(root, textvariable=Name3, width=14, bd=2, relief="sunken", font="Montserrat 14", fg="#424949")
Name_box3.place(x=435, y=344)


Phone = StringVar()
Phone_box = Entry(root, textvariable=Phone, width=40, bd=2, relief="sunken", font="Montserrat 14", fg="#424949")
Phone_box.place(x=695, y=342)

Year = Combobox(obj, values=['1st Year', '2nd Year', '3rd Year', '4th Year'], font="Montserrat 14", width=40, foreground="#424949", state='r')
Year.place(x=658, y=190)
Year.set("Please Select")

Degree = Combobox(obj, values=['B.Tech. in Computer Science and Engineering', 'B.Tech. in Data Science Engineering', 'B.Tech. in Data Science Engineering', 'B.Tech. in Information Technology', 'B.Tech. in Computer Science and Engineering (Cyber Security)', 'B.Tech. in Computer Science and Engineering (Artificial Intelligence and Machine Learning)', 'B.Tech. in Computer Science and Engineering (IoT)', 'B.Tech. in Civil Engineering', 'B.Tech. in Mechanical Engineering', 'B.Tech. in Electrical Engineering', 'B.Tech. in Electronics and Telecommunication Engineering', 'B.Tech. in Electronics Engineering', 'B.Tech. in Computer Science and Engineering (Artificial Intelligence)', 'B.Tech. in Computer Science and Engineering (Internet of Things)'], font="Montserrat 14", width=47, foreground="#424949",  state='r')
Degree.place(x=58, y=190)
Degree.set("Please Select")

Email = StringVar()
Email_box = Entry(root, textvariable=Email, width=44, bd=2, relief="sunken", font="Montserrat 14", fg="#424949")
Email_box.place(x=95, y=580)

DOB1 = Combobox(obj, values=['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December'], font="Montserrat 14", width=12, foreground="#424949", state='r')
DOB1.place(x=650, y=298)
DOB1.set("Select a Month")

DOB2 = Combobox(obj, values=['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13', '14', '15', '16', '17', '18', '19', '20', '21', '22', '23', '24', '25', '26', '27', '28', '29', '30', '31'], font="Montserrat 14", width=12, foreground="#424949", state='r')
DOB2.place(x=810, y=298)
DOB2.set("Select a Day")

DOB3 = Combobox(obj, values=['1990', '1991', '1992', '1993', '1994', '1995', '1996', '1997', '1998', '1999', '2000','2001', '2002', '2003', '2004', '2005', '2006', '2007', '2008', '2009', '2010', '2011', '2012', '2013', '2014', '2015', '2016', '2017', '2018', '2019', '2020', '2021', '2022', '2023', '2024'], font="Montserrat 14", width=12, foreground="#424949", state='r')
DOB3.place(x=970, y=298)
DOB3.set("Select a Year")

Address = StringVar()
Address_box = Entry(root, textvariable=Address, width=44, bd=2, relief="sunken", font="Montserrat 14", fg="#424949")
Address_box.place(x=95, y=697)

Gender = IntVar()
G1 = Radiobutton(obj, text="Male", font="Montserrat 14", variable=Gender, value=1, command=selection, fg="#212F3C", bg="white")
G1.place(x=705, y=415)
G2 = Radiobutton(obj, text="Female", font="Montserrat 14", variable=Gender, value=2, command=selection, fg="#212F3C", bg="white")
G2.place(x=785, y=415)

# Profile Photo
f = Frame(root, bd=2, width=200, height=200, bg="white", relief="groove")
f.place(x=1270, y=250)
profile = (PhotoImage(file="Images/user1.png"))
p = Label(f, bg="white", image=profile)
p.place(x=0, y=0)

# Buttons
Button(root, text="Upload", width=16, font="Montserrat 14 bold", bg="white", command=Showimage).place(x=1270, y=490)
SaveButton = Button(root, text="Save", width=16, font="Montserrat 14 bold", bg="white", command=Save)
SaveButton.place(x=1270, y=550)
Button(root, text="Reset", width=16, font="Montserrat 14 bold", bg="white", command=Clear).place(x=1270, y=610)
Button(root, text="Exit", width=16, font="Montserrat 14 bold", bg="white", command=Exit).place(x=1270, y=670)

root.mainloop()